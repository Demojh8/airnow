from openpyxl import load_workbook

class ExcelReader:  

    def __init__(self):
        self.data = []

    def loadData(self, filename, min_row_val, max_col_val, max_row_val):
        wb = load_workbook(filename)
        sheet_ranges = wb['Sheet1']
        
        for row in sheet_ranges.iter_rows(min_row=min_row_val, max_col=max_col_val, max_row=max_row_val):
            for cell in row:
                if(cell.value != None):
                    self.data.append(cell.value)                    

        return
